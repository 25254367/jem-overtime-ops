"""Build the frozen hand-labelling sample for the note-classification check.

Two strata so the check is honest AND the rare categories get enough cases:
  - RANDOM  (~100): representative -> honest overall accuracy
  - TARGETED (~80): oversamples multilingual, short/ambiguous, rare categories,
                    and likely "notes that lie" -> enough cases to measure the
                    hard categories and failure modes

Outputs:
  check/gold/gold_sample.xlsx   <- fill this (dropdowns built in)
  check/gold/gold_sample.csv    <- same content, for anyone without Excel
  check/gold/gold_sample_strata.csv  <- hidden sampling metadata, for scoring later

Reproducible: fixed seed. Once gold_sample.csv exists it is the frozen sample
and is never re-selected -- re-running only rebuilds the xlsx from it.
"""
from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd

from .load import load_export

SEED = 20260901
N_RANDOM = 100
N_TARGETED = 80
OUT = Path("check/gold/gold_sample.csv")
XLSX = Path("check/gold/gold_sample.xlsx")

# the labelling columns and their allowed values (blank = free text)
LABEL_OPTIONS = [
    "client_requested", "absence_cover", "late_handover",
    "equipment_failure", "routine", "blank", "unclassified",
]
ABSENCE_REASON_OPTIONS = ["no_show", "sick", "leave", "unknown"]
AUTHORISED_OPTIONS = ["yes", "unclear"]
CONFIDENCE_OPTIONS = ["confident", "ambiguous"]
FILL_COLUMNS = ["label", "absence_reason", "authorised", "absentee",
                "confidence", "comment"]

# rough keyword pre-tag -- ONLY used to steer stratified sampling, never shown
# to the labeller and never used as an answer.
_AFR = re.compile(r"\b(nie|niks|gedek|geddek|siek|gemeld|oorhandiging|sleutels|"
                  r"aflos|opgedaag|aanbly|klient|klent|ekstra|ure|masjien|"
                  r"stukkend|gewag|nagskof|geen)\b", re.I)
_ZUL = re.compile(r"\b(akezanga|namhlanje|ngimele|yena|yean|akukho|lutho|"
                  r"ngicela|umsebenzi|wami)\b", re.I)
_CLIENT = re.compile(r"client|klient|klent|centre m|site manager|requested by|"
                     r"approved|signed off|sign off|ok'd|stocktake|deep clean|"
                     r"load in|delivery|event|audit|patrol per client", re.I)
_ABSENCE = re.compile(r"no show|no-show|didn.?t|akezanga|opgedaag|booked off|"
                      r"off sick|siek|absent|covering|covered|stood in|relief|"
                      r"aflos|no replacement|2 posts|took .*(shift|rounds|post)|"
                      r"double (shift|duty)|clinic|family responsib|next shift|"
                      r"nobody came|no relief", re.I)
_HANDOVER = re.compile(r"handover|oorhandiging|keys|ob book|paperwork|"
                       r"waited .*min|delayed by", re.I)
_EQUIP = re.compile(r"machine|macine|machnie|generator|scrubber|buffer|lift|"
                    r"gate motor|masjien|stukkend|kaput|broke|broken|fault|"
                    r"down again|load ?shed|power", re.I)
_ROUTINE = re.compile(r"quiet|nothing to report|no incidents?|all (quiet|good|"
                      r"fine)|as per normal|akukho lutho|niks om te rap|"
                      r"no issues", re.I)


def _rough_tag(t: str) -> str:
    s = str(t).strip().lower()
    if not s or s in {"-", ".", "n/a", "ntr", "ok", "sharp", "fine", "all good"}:
        return "blank"
    if _CLIENT.search(s) and _ABSENCE.search(s):
        return "conflict"          # candidate "note that lies"
    if _EQUIP.search(s):
        return "equipment_failure"
    if _ABSENCE.search(s):
        return "absence_cover"
    if _HANDOVER.search(s):
        return "late_handover"
    if _CLIENT.search(s):
        return "client_requested"
    if _ROUTINE.search(s):
        return "routine"
    return "unclassified"


def _lang(t: str) -> str:
    s = str(t).lower()
    if _ZUL.search(s):
        return "zu"
    if _AFR.search(s):
        return "af"
    return "en"


def _len_bucket(n: int) -> str:
    for hi, name in [(4, "xs"), (16, "s"), (31, "m"), (46, "l")]:
        if n < hi:
            return name
    return "xl"


def build(data_dir: str = "data") -> pd.DataFrame:
    if OUT.exists():
        frozen = pd.read_csv(OUT, dtype=str).fillna("")
        n_labelled = int((frozen.get("label", pd.Series(dtype=str))
                          .astype(str).str.strip() != "").sum())
        if n_labelled > 0:
            print(f"{OUT} exists with {n_labelled} labels — returning the "
                  f"frozen sample. Delete it to start over.")
            if not XLSX.exists():
                _write_xlsx(frozen.reindex(
                    columns=["row", "shift_id", "logged_by", "note", *FILL_COLUMNS],
                    fill_value=""))
            return frozen
        # nothing labelled yet -> safe to rebuild the sheet (same 180 shift_ids,
        # current columns) without losing any work.
        print(f"{OUT} exists but is unlabelled — rebuilding sheet & xlsx.")
        sheet = frozen[["row", "shift_id", "logged_by", "note"]].copy()
        for col in FILL_COLUMNS:
            sheet[col] = ""
        OUT.parent.mkdir(exist_ok=True)
        sheet.to_csv(OUT, index=False)
        _write_xlsx(sheet)
        print(f"rebuilt {len(sheet)} rows -> {XLSX} (+ .csv)")
        return sheet

    exp = load_export(data_dir)
    notes = exp.shift_notes.copy()
    notes["note"] = notes["note"].fillna("")
    notes["rough_tag"] = notes["note"].map(_rough_tag)
    notes["lang"] = notes["note"].map(_lang)
    notes["len_bucket"] = notes["note"].str.len().map(_len_bucket)

    rng = np.random.default_rng(SEED)

    # ---- RANDOM stratum ----
    random_idx = rng.choice(notes.index, size=N_RANDOM, replace=False)
    rand = notes.loc[random_idx].assign(stratum="random")

    pool = notes.drop(index=random_idx)

    # ---- TARGETED stratum: fill quotas from the rare/hard buckets ----
    quotas = {
        ("conflict", None): 12,          # notes that lie
        ("equipment_failure", None): 12,
        ("late_handover", None): 10,
        ("unclassified", None): 12,
        (None, "zu"): 12,                # isiZulu
        (None, "af"): 12,               # Afrikaans
        ("absence_cover", None): 10,     # boundary cases
    }
    picked: list = []
    for (tag, lang), k in quotas.items():
        sub = pool
        if tag is not None:
            sub = sub[sub["rough_tag"] == tag]
        if lang is not None:
            sub = sub[sub["lang"] == lang]
        sub = sub.drop(index=[i for i in picked if i in sub.index], errors="ignore")
        take = min(k, len(sub))
        if take:
            chosen = rng.choice(sub.index, size=take, replace=False)
            picked.extend(chosen.tolist())
    picked = list(dict.fromkeys(picked))[:N_TARGETED]
    # top up to N_TARGETED with random draws from the remaining pool if short
    if len(picked) < N_TARGETED:
        rest = pool.drop(index=picked, errors="ignore")
        extra = rng.choice(rest.index, size=N_TARGETED - len(picked), replace=False)
        picked.extend(extra.tolist())
    tgt = notes.loc[picked].assign(stratum="targeted")

    sample = pd.concat([rand, tgt]).sample(frac=1, random_state=SEED)  # shuffle
    sample = sample.reset_index(drop=True)
    sample.insert(0, "row", range(1, len(sample) + 1))

    # the labelling worksheet: only what shift_notes.csv gives you (shift_id,
    # logged_by, note) so the gold labels are made under the same information
    # the text classifier sees. BLANK columns to fill.
    sheet = sample[["row", "shift_id", "logged_by", "note"]].copy()
    for col in FILL_COLUMNS:
        sheet[col] = ""

    OUT.parent.mkdir(exist_ok=True)
    sheet.to_csv(OUT, index=False)
    sample[["row", "shift_id", "stratum", "rough_tag", "lang", "len_bucket"]].to_csv(
        OUT.with_name("gold_sample_strata.csv"), index=False
    )
    _write_xlsx(sheet)
    print(f"wrote {len(sheet)} rows -> {XLSX} (+ .csv)")
    print(f"  strata: {sample['stratum'].value_counts().to_dict()}")
    print(f"  languages: {sample['lang'].value_counts().to_dict()}")
    print(f"  rough tags: {sample['rough_tag'].value_counts().to_dict()}")
    return sheet


def _write_xlsx(sheet: pd.DataFrame) -> None:
    """Same content as the csv, with dropdown data-validation on the label
    columns and a frozen header. Fill it in Excel / Numbers / Sheets."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "gold_sample"
    ws.append(list(sheet.columns))
    for _, r in sheet.iterrows():
        ws.append(list(r))

    ws.freeze_panes = "A2"
    widths = {"row": 5, "shift_id": 10, "logged_by": 9, "note": 70,
              "label": 18, "absence_reason": 14, "authorised": 11,
              "absentee": 14, "confidence": 12, "comment": 40}
    col_idx = {c: i + 1 for i, c in enumerate(sheet.columns)}
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx[name])].width = w

    last = len(sheet) + 1
    for name, options in [
        ("label", LABEL_OPTIONS),
        ("absence_reason", ABSENCE_REASON_OPTIONS),
        ("authorised", AUTHORISED_OPTIONS),
        ("confidence", CONFIDENCE_OPTIONS),
    ]:
        letter = get_column_letter(col_idx[name])
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,
            showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
        )
        dv.error = "Pick a value from the list"
        dv.prompt = " / ".join(options)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last}")

    wb.save(XLSX)


if __name__ == "__main__":
    import sys

    build(sys.argv[1] if len(sys.argv) > 1 else "data")
